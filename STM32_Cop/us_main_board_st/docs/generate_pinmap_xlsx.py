#!/usr/bin/env python3
"""
STM32G474RC 핀맵 Excel 생성 스크립트
4개 시트: 기능별 핀맵, 물리핀 순서, PCB 배치 가이드, 핀 변경 이력
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 공용 스타일 ──
FONT_TITLE = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
FONT_GROUP = Font(name="맑은 고딕", size=10, bold=True, color="1F4E79")
FONT_NORMAL = Font(name="맑은 고딕", size=10)
FONT_WARN = Font(name="맑은 고딕", size=10, bold=True, color="CC0000")
FONT_OK = Font(name="맑은 고딕", size=10, bold=True, color="006600")
FONT_SPARE = Font(name="맑은 고딕", size=10, color="888888")
FONT_NOTE = Font(name="맑은 고딕", size=9, italic=True, color="555555")

FILL_TITLE = PatternFill("solid", fgColor="1F4E79")
FILL_HEADER = PatternFill("solid", fgColor="2E75B6")
FILL_GROUP = PatternFill("solid", fgColor="D6E4F0")  # 밝은 파랑
FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
FILL_ZEBRA = PatternFill("solid", fgColor="F2F7FB")
FILL_WARN = PatternFill("solid", fgColor="FFF0F0")
FILL_OK = PatternFill("solid", fgColor="F0FFF0")
FILL_SPARE = PatternFill("solid", fgColor="F5F5F5")
FILL_ANALOG = PatternFill("solid", fgColor="FFFDE7")  # 연한 노랑 (아날로그)
FILL_PWM = PatternFill("solid", fgColor="E8F5E9")     # 연한 초록 (PWM)
FILL_COMM = PatternFill("solid", fgColor="E3F2FD")    # 연한 파랑 (통신)
FILL_LCD = PatternFill("solid", fgColor="F3E5F5")     # 연한 보라 (LCD)
FILL_BTN = PatternFill("solid", fgColor="FFF3E0")     # 연한 주황 (버튼)
FILL_EXT = PatternFill("solid", fgColor="E0F7FA")     # 연한 청록 (외부제어)
FILL_DBG = PatternFill("solid", fgColor="ECEFF1")     # 연한 회색 (디버그)
FILL_PWR = PatternFill("solid", fgColor="FAFAFA")     # 옅은 회색 (전원)
FILL_INTERNAL = PatternFill("solid", fgColor="FCE4EC") # 연한 핑크 (내부 아날로그)

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def apply_style(ws, row, col, font=FONT_NORMAL, fill=FILL_WHITE, alignment=ALIGN_CENTER):
    cell = ws.cell(row=row, column=col)
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = THIN_BORDER
    return cell


def write_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        cell = apply_style(ws, row, col_start + i, FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
        cell.value = h


def write_data_row(ws, row, data, col_start=1, font=FONT_NORMAL, fill=None, alignment=ALIGN_CENTER):
    for i, val in enumerate(data):
        f = fill if fill else (FILL_ZEBRA if row % 2 == 0 else FILL_WHITE)
        cell = apply_style(ws, row, col_start + i, font, f, alignment)
        cell.value = val


def write_title(ws, row, title, col_count):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = apply_style(ws, row, 1, FONT_TITLE, FILL_TITLE, ALIGN_CENTER)
    cell.value = title
    ws.row_dimensions[row].height = 30


def write_section_header(ws, row, text, col_count):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = apply_style(ws, row, 1, FONT_GROUP, FILL_GROUP, ALIGN_LEFT)
    cell.value = text
    ws.row_dimensions[row].height = 22


def auto_width(ws, col_widths):
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w


# ═══════════════════════════════════════════════════════════
# 시트 1: 기능별 핀맵
# ═══════════════════════════════════════════════════════════
def create_sheet_functional(wb):
    ws = wb.active
    ws.title = "기능별 핀맵"
    ws.sheet_properties.tabColor = "2E75B6"
    col_count = 6
    auto_width(ws, [8, 8, 16, 48, 14, 14])

    write_title(ws, 1, "STM32G474RCT6 기능별 핀맵 (LQFP64)", col_count)
    r = 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_count)
    cell = apply_style(ws, r, 1, FONT_NOTE, FILL_WHITE, ALIGN_LEFT)
    cell.value = "문서 v2.0 | 최종 수정: 2026-03-30 | copilot-instructions.md 동기화"
    r = 3

    groups = [
        ("🔊 초음파 발진 (TIM1 고급 타이머)", FILL_PWM, [
            ["PA8",  35, "TIM1_CH1 (AF6)",  "초음파 PWM 출력 — 하이사이드 게이트 드라이버(HIN) 입력", "Output", "Very High"],
            ["PB13", 28, "TIM1_CH1N (AF6)", "초음파 PWM 상보 출력 — 로우사이드 게이트 드라이버(LIN) 입력", "Output", "Very High"],
        ]),
        ("⚡ 위상제어 조광기 출력 (TIM3 → ATtiny85)", FILL_PWM, [
            ["PB0", 22, "TIM3_CH3 (AF2)", "ATtiny85 위상제어 듀티 지령 전달용 3kHz PWM 출력", "Output", "Medium"],
        ]),
        ("📡 내장 아날로그 — 비교기(COMP) + DAC3 + TIM2 캡처", FILL_INTERNAL, [
            ["PA1",  13, "COMP1_INP",              "전압(V) 위상 제로크로싱 검출 — 비교기 양(+)입력 【전용, ADC 비공유】", "Analog", "—"],
            ["PA7",  19, "COMP2_INP",              "전류(I) 위상 제로크로싱 검출 — 비교기 양(+)입력 (CT센서) 【전용, GPIO 비공유】", "Analog", "—"],
            ["—",    "—", "DAC3_CH1 → COMP1_INM",  "COMP1 반전입력 기준전압 (~1.65V = VDDA/2) 【내부 크로스바, 외부핀 없음】", "Internal", "—"],
            ["—",    "—", "DAC3_CH2 → COMP2_INM",  "COMP2 반전입력 기준전압 (~1.65V = VDDA/2) 【내부 크로스바, 외부핀 없음】", "Internal", "—"],
            ["—",    "—", "COMP1_OUT → TIM2_IC1",  "전압 제로크로싱 타임스탬프 캡처 【내부 라우팅】", "Internal", "—"],
            ["—",    "—", "COMP2_OUT → TIM2_IC2",  "전류 제로크로싱 타임스탬프 캡처 【내부 라우팅】", "Internal", "—"],
        ]),
        ("🎛️ ADC 입력 (가변저항 + 전류 센서)", FILL_ANALOG, [
            ["PA0",  12, "ADC1_IN1",  "보드 내장 가변저항 전압 (초음파 출력/위상 레벨)", "Analog", "—"],
            ["PB11", 26, "ADC1_IN14", "외부 판넬 가변저항 전압 (외부 출력/위상 레벨 지정)", "Analog", "—"],
            ["PA4",  16, "ADC2_IN17", "AC 전원 입력단 시스템 소모 전류 측정용 (ACS722/TMCS1100)", "Analog", "—"],
        ]),
        ("📺 LCD1602 (4비트 병렬 모드)", FILL_LCD, [
            ["PB12", 27, "GPIO_Output", "LCD RS (Register Select)", "Output", "Low"],
            ["PB14", 29, "GPIO_Output", "LCD EN (Enable)", "Output", "Low"],
            ["PB15", 30, "GPIO_Output", "LCD D4", "Output", "Low"],
            ["PC6",  31, "GPIO_Output", "LCD D5", "Output", "Low"],
            ["PC7",  32, "GPIO_Output", "LCD D6", "Output", "Low"],
            ["PC8",  33, "GPIO_Output", "LCD D7", "Output", "Low"],
        ]),
        ("🔘 택트 스위치 (내부 풀업, Active LOW)", FILL_BTN, [
            ["PC0", 8,  "GPIO_Input (PU)", "BTN_MENU — 메뉴 진입 및 상위 메뉴 복귀", "Input", "—"],
            ["PC1", 9,  "GPIO_Input (PU)", "BTN_UP — 메뉴 위로 / 값 증가", "Input", "—"],
            ["PC2", 10, "GPIO_Input (PU)", "BTN_DOWN — 메뉴 아래로 / 값 감소", "Input", "—"],
            ["PC3", 11, "GPIO_Input (PU)", "BTN_OK — 선택/확인 및 초음파 Start/Stop 겸용", "Input", "—"],
        ]),
        ("🔌 Modbus RS485 (USART2 + 방향 제어)", FILL_COMM, [
            ["PA2",  14, "USART2_TX (AF7)",    "RS485 TXD", "AF Output", "Very High"],
            ["PA3",  15, "USART2_RX (AF7)",    "RS485 RXD", "AF Input", "—"],
            ["PA5",  17, "GPIO_Output",         "RS485 DE (Driver Enable, HIGH=송신) — USART2 인접 배치", "Output", "High"],
            ["PA6",  18, "GPIO_Output",         "RS485 /RE (Receiver Enable, LOW=수신) — USART2 인접 배치", "Output", "High"],
            ["PC11", 46, "GPIO_Output",         "RS485 RTERM (종단 저항 120Ω ON/OFF 제어)", "Output", "Low"],
            ["PC12", 47, "GPIO_Input",          "Board_Detect (RS485 옵션 보드 장착 인식, 메뉴 연동)", "Input", "—"],
        ]),
        ("🔧 외부 제어 인터페이스 (스위치 입력 + 릴레이/부저 출력)", FILL_EXT, [
            ["PC4",  20, "GPIO_Output", "부저 출력 — 버튼 클릭음 + 에러 알람 겸용", "Output", "Low"],
            ["PC5",  21, "GPIO_Output", "End Signal 릴레이 — 사이클 종료 후 0.5초 ON", "Output", "Low"],
            ["PA12", 39, "GPIO_Output", "Fault 릴레이 — 에러 발생 상태 지시", "Output", "Low"],
            ["PA15", 44, "GPIO_Output", "Running 릴레이 — 정상 동작 상태 지시 (G4는 SWD 기본, 바로 사용)", "Output", "Low"],
            ["PB1",  23, "GPIO_Input",  "Remote ON/OFF — 외부 기기 리모트 제어 입력", "Input", "—"],
            ["PB2",  24, "GPIO_Input",  "Sweep 스위치 — 수동 스윕 분산 기능 푸쉬락", "Input", "—"],
            ["PB10", 25, "GPIO_Input",  "Timer 스위치 — 내장 타이머 사용 유무 푸쉬락", "Input", "—"],
        ]),
        ("🐛 디버그 / 시스템", FILL_DBG, [
            ["PA9",  36, "USART1_TX (AF7)", "디버그 메시지용 UART TX (115200bps, 선택사항)", "AF Output", "High"],
            ["PA10", 37, "USART1_RX (AF7)", "디버그 메시지 수신용 UART RX (선택사항)", "AF Input", "—"],
            ["PA13", 40, "SWDIO",           "SWD 데이터 입출력 — GPIO 사용 금지 ⛔", "SWD", "—"],
            ["PA14", 43, "SWCLK",           "SWD 클럭 입력 — GPIO 사용 금지 ⛔", "SWD", "—"],
            ["PB3",  49, "SWO",             "Serial Wire Output — Trace용 (선택사항, 보통 생략)", "SWD", "—"],
            ["PC10", 45, "GPIO_Output",     "상태 LED — Run/Error 토글 표시", "Output", "Low"],
            ["NRST",  7, "RESET",           "MCU 리셋 (ST-Link 연결 강력 권장)", "System", "—"],
            ["BOOT0",54, "—",               "부트 모드 선택 (HIGH=부트로더, 점퍼/버튼 권장)", "System", "—"],
        ]),
        ("📦 I2C 확장 (EEPROM 등, 선택사항)", FILL_COMM, [
            ["PB8", 55, "I2C1_SCL (AF4)", "I2C EEPROM SCL (파라미터 저장 확장용)", "AF OD", "High"],
            ["PB9", 56, "I2C1_SDA (AF4)", "I2C EEPROM SDA (파라미터 저장 확장용)", "AF OD", "High"],
        ]),
        ("🔓 여유(Spare) 핀", FILL_SPARE, [
            ["PC9",  34, "—", "GPIO 자유 사용 (RS485 DE 이전 핀)", "—", "—"],
            ["PA11", 38, "—", "TIM1_CH4(AF11) 또는 GPIO", "—", "—"],
            ["PB4",  50, "—", "GPIO 또는 TIM16_CH1(AF1)", "—", "—"],
            ["PB5",  51, "—", "GPIO 또는 SPI1_MOSI(AF5)", "—", "—"],
            ["PB6",  52, "—", "I2C1_SCL(AF4) 확장 또는 GPIO", "—", "—"],
            ["PB7",  53, "—", "I2C1_SDA(AF4) 확장 또는 GPIO", "—", "—"],
            ["PD2",  48, "—", "TIM3_ETR(AF2) 또는 GPIO", "—", "—"],
        ]),
    ]

    headers = ["핀", "물리핀#", "AF / 모드", "기능 설명", "I/O 방향", "GPIO 속도"]
    r += 1

    for group_name, group_fill, rows in groups:
        write_section_header(ws, r, group_name, col_count)
        r += 1
        write_header_row(ws, r, headers)
        r += 1
        for rd in rows:
            for i, val in enumerate(rd):
                if rd[0] == "—":
                    f = FILL_INTERNAL
                    fn = FONT_NOTE
                else:
                    f = group_fill if r % 2 == 0 else FILL_WHITE
                    fn = FONT_NORMAL
                cell = apply_style(ws, r, i+1, fn, f, ALIGN_CENTER if i != 3 else ALIGN_LEFT)
                cell.value = val
            r += 1
        r += 1  # 그룹 간 빈 줄

    # DAC 충돌 섹션
    write_section_header(ws, r, "⚠️ DAC 핀 충돌 상태 (LQFP64)", col_count)
    r += 1
    dac_headers = ["DAC", "채널", "출력 핀/경로", "충돌 대상", "상태", "비고"]
    write_header_row(ws, r, dac_headers)
    r += 1
    dac_data = [
        ["DAC3", "CH1", "→ COMP1_INM (내부)", "없음 (내부 전용)", "✅ 사용", "기준전압 공급"],
        ["DAC3", "CH2", "→ COMP2_INM (내부)", "없음 (내부 전용)", "✅ 사용", "기준전압 공급"],
        ["DAC4", "CH1/CH2", "내부 전용 (예비)", "없음", "🔵 예비", "COMP5~7 확장용"],
        ["DAC1", "CH1", "PA4", "ADC2_IN17 (전류센서)", "🔴 충돌", "사용 불가"],
        ["DAC1", "CH2", "PA5", "RS485 DE", "🔴 충돌", "사용 불가"],
        ["DAC2", "CH1", "PA6", "RS485 /RE", "🔴 충돌", "사용 불가"],
    ]
    for rd in dac_data:
        fill = FILL_OK if "✅" in str(rd[4]) else (FILL_WARN if "🔴" in str(rd[4]) else FILL_WHITE)
        fn = FONT_OK if "✅" in str(rd[4]) else (FONT_WARN if "🔴" in str(rd[4]) else FONT_NORMAL)
        write_data_row(ws, r, rd, font=fn, fill=fill)
        r += 1
    r += 1

    # OPAMP 충돌 섹션
    write_section_header(ws, r, "⚠️ OPAMP 핀 충돌 상태 (LQFP64 — 전부 사용 불가)", col_count)
    r += 1
    oa_headers = ["OPAMP", "VINP 핀", "VOUT 핀", "충돌 대상", "상태", "비고"]
    write_header_row(ws, r, oa_headers)
    r += 1
    oa_data = [
        ["OPAMP1", "PA1, PA3, PA7", "PA2", "COMP1, USART2_RX/TX", "🔴 사용불가", ""],
        ["OPAMP2", "PA7", "PA6", "COMP2, RS485 /RE", "🔴 사용불가", ""],
        ["OPAMP3", "PB0, PB13, PA1", "PB1", "TIM3_CH3, TIM1_CH1N, Remote", "🔴 사용불가", ""],
        ["OPAMP4", "PB13, PB11", "PB12", "TIM1_CH1N, 외부VR, LCD RS", "🔴 사용불가", ""],
        ["OPAMP5", "PB14, PC3", "PA8", "LCD EN, BTN_OK, TIM1_CH1", "🔴 사용불가", ""],
        ["OPAMP6", "PB12", "PB11", "LCD RS, 외부VR ADC", "🔴 사용불가", ""],
    ]
    for rd in oa_data:
        write_data_row(ws, r, rd, font=FONT_WARN, fill=FILL_WARN)
        r += 1

    # 노트
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_count)
    cell = apply_style(ws, r, 1, FONT_NOTE, FILL_WHITE, ALIGN_LEFT)
    cell.value = "※ OPAMP: LQFP100 이상 마이그레이션 시 일부 활용 여지 있음. CT센서 신호 컨디셔닝은 외부 회로(Burden저항+바이어스+클램핑)로 해결."
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_count)
    cell = apply_style(ws, r, 1, FONT_NOTE, FILL_WHITE, ALIGN_LEFT)
    cell.value = "※ DAC3 = 내부 전용 DAC (외부 핀 없음). 칩 내부 크로스바로 COMP1~4 INM에 직결. 소프트웨어로 mV 단위 기준전압을 실시간 보정 가능."

    ws.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════
# 시트 2: 물리핀 순서
# ═══════════════════════════════════════════════════════════
def create_sheet_physical(wb):
    ws = wb.create_sheet("물리핀 순서")
    ws.sheet_properties.tabColor = "00B050"
    col_count = 7
    auto_width(ws, [10, 8, 8, 16, 40, 14, 16])

    write_title(ws, 1, "LQFP64 물리 핀 번호 순서 (Pin 1 ~ 64)", col_count)
    r = 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_count)
    cell = apply_style(ws, r, 1, FONT_NOTE, FILL_WHITE, ALIGN_LEFT)
    cell.value = "색상 범례: 🟡아날로그  🟢PWM  🔵통신  🟣LCD  🟠버튼  🩵외부제어  ⬜디버그/전원  🩷내부아날로그"

    sides = [
        ("◀ Left Side (Pin 1 ~ 16)", [
            [1,  "VBAT", "—",           "Power",          "배터리 백업 (미사용 시 VDD 연결)", "전원", FILL_PWR],
            [2,  "PC13", "—",           "—",              "(미사용)", "—", FILL_PWR],
            [3,  "PC14", "—",           "—",              "(미사용, OSC32_IN)", "—", FILL_PWR],
            [4,  "PC15", "—",           "—",              "(미사용, OSC32_OUT)", "—", FILL_PWR],
            [5,  "PH0",  "—",           "—",              "OSC_IN (HSE 사용 시) 또는 미사용", "클럭", FILL_PWR],
            [6,  "PH1",  "—",           "—",              "OSC_OUT (HSE 사용 시) 또는 미사용", "클럭", FILL_PWR],
            [7,  "NRST", "—",           "RESET",          "MCU 리셋 (ST-Link 연결 강력 권장)", "디버그", FILL_DBG],
            [8,  "PC0",  "—",           "GPIO_Input (PU)","BTN_MENU — 메뉴 진입/복귀", "택트 스위치", FILL_BTN],
            [9,  "PC1",  "—",           "GPIO_Input (PU)","BTN_UP — 위로/값 증가", "택트 스위치", FILL_BTN],
            [10, "PC2",  "—",           "GPIO_Input (PU)","BTN_DOWN — 아래로/값 감소", "택트 스위치", FILL_BTN],
            [11, "PC3",  "—",           "GPIO_Input (PU)","BTN_OK — 선택/확인, Start/Stop", "택트 스위치", FILL_BTN],
            [12, "PA0",  "ADC1_IN1",    "ADC Analog",     "보드 내장 가변저항 (출력/위상 레벨)", "ADC", FILL_ANALOG],
            [13, "PA1",  "COMP1_INP",   "Analog",         "전압(V) 위상 제로크로싱 검출 비교기 (+)", "COMP", FILL_INTERNAL],
            [14, "PA2",  "USART2_TX",   "AF7",            "RS485 TXD", "RS485", FILL_COMM],
            [15, "PA3",  "USART2_RX",   "AF7",            "RS485 RXD", "RS485", FILL_COMM],
            [16, "PA4",  "ADC2_IN17",   "ADC Analog",     "AC 전류 센서 (ACS722/TMCS1100)", "ADC", FILL_ANALOG],
        ]),
        ("▼ Bottom Side (Pin 17 ~ 32)", [
            [17, "PA5",  "—",            "GPIO_Output",   "RS485 DE (HIGH=송신) — USART2 인접", "RS485", FILL_COMM],
            [18, "PA6",  "—",            "GPIO_Output",   "RS485 /RE (LOW=수신) — USART2 인접", "RS485", FILL_COMM],
            [19, "PA7",  "COMP2_INP",    "Analog",        "전류(I) 위상 제로크로싱 비교기 (+) CT센서", "COMP", FILL_INTERNAL],
            [20, "PC4",  "—",            "GPIO_Output",   "부저 (조작음 + 에러 알람 겸용)", "외부 제어", FILL_EXT],
            [21, "PC5",  "—",            "GPIO_Output",   "End Signal 릴레이 (사이클 종료 0.5s)", "외부 제어", FILL_EXT],
            [22, "PB0",  "TIM3_CH3",     "AF2",           "ATtiny85 3kHz PWM (위상제어 지령)", "위상제어", FILL_PWM],
            [23, "PB1",  "—",            "GPIO_Input",    "Remote ON/OFF (외부 리모트 제어)", "외부 제어", FILL_EXT],
            [24, "PB2",  "—",            "GPIO_Input",    "Sweep 스위치 (수동 스윕 푸쉬락)", "외부 제어", FILL_EXT],
            [25, "PB10", "—",            "GPIO_Input",    "Timer 스위치 (내장 타이머 푸쉬락)", "외부 제어", FILL_EXT],
            [26, "PB11", "ADC1_IN14",    "ADC Analog",    "외부 판넬 가변저항 (외부 출력 레벨)", "ADC", FILL_ANALOG],
            [27, "PB12", "—",            "GPIO_Output",   "LCD RS (Register Select)", "LCD", FILL_LCD],
            [28, "PB13", "TIM1_CH1N",    "AF6",           "초음파 PWM 상보 출력 (로우사이드 FET)", "초음파", FILL_PWM],
            [29, "PB14", "—",            "GPIO_Output",   "LCD EN (Enable)", "LCD", FILL_LCD],
            [30, "PB15", "—",            "GPIO_Output",   "LCD D4", "LCD", FILL_LCD],
            [31, "PC6",  "—",            "GPIO_Output",   "LCD D5", "LCD", FILL_LCD],
            [32, "PC7",  "—",            "GPIO_Output",   "LCD D6", "LCD", FILL_LCD],
        ]),
        ("▶ Right Side (Pin 33 ~ 48)", [
            [33, "PC8",  "—",           "GPIO_Output",    "LCD D7", "LCD", FILL_LCD],
            [34, "PC9",  "—",           "—",              "(Spare) GPIO 자유 사용", "여유", FILL_SPARE],
            [35, "PA8",  "TIM1_CH1",    "AF6",            "초음파 PWM 출력 (하이사이드 FET)", "초음파", FILL_PWM],
            [36, "PA9",  "USART1_TX",   "AF7",            "디버그 UART TX (115200bps)", "디버그", FILL_DBG],
            [37, "PA10", "USART1_RX",   "AF7",            "디버그 UART RX", "디버그", FILL_DBG],
            [38, "PA11", "—",           "—",              "(Spare) TIM1_CH4(AF11) 또는 GPIO", "여유", FILL_SPARE],
            [39, "PA12", "—",           "GPIO_Output",    "Fault 릴레이 (에러 상태 지시)", "외부 제어", FILL_EXT],
            [40, "PA13", "SWDIO",       "SWD",            "SWD 데이터 ⛔ GPIO 사용 금지", "디버그", FILL_DBG],
            [41, "VSS",  "—",           "Power",          "GND", "전원", FILL_PWR],
            [42, "VDD",  "—",           "Power",          "3.3V", "전원", FILL_PWR],
            [43, "PA14", "SWCLK",       "SWD",            "SWD 클럭 ⛔ GPIO 사용 금지", "디버그", FILL_DBG],
            [44, "PA15", "—",           "GPIO_Output",    "Running 릴레이 (정상 동작 지시)", "외부 제어", FILL_EXT],
            [45, "PC10", "—",           "GPIO_Output",    "상태 LED (Run/Error 토글)", "디버그", FILL_DBG],
            [46, "PC11", "—",           "GPIO_Output",    "RS485 RTERM (종단 저항 ON/OFF)", "RS485", FILL_COMM],
            [47, "PC12", "—",           "GPIO_Input",     "Board_Detect (RS485 옵션보드 인식)", "RS485", FILL_COMM],
            [48, "PD2",  "—",           "—",              "(Spare) TIM3_ETR(AF2) 또는 GPIO", "여유", FILL_SPARE],
        ]),
        ("▲ Top Side (Pin 49 ~ 64)", [
            [49, "PB3",  "SWO",         "SWD",            "Serial Wire Output (선택사항)", "디버그", FILL_DBG],
            [50, "PB4",  "—",           "—",              "(Spare) GPIO 또는 TIM16_CH1(AF1)", "여유", FILL_SPARE],
            [51, "PB5",  "—",           "—",              "(Spare) GPIO 또는 SPI1_MOSI(AF5)", "여유", FILL_SPARE],
            [52, "PB6",  "—",           "—",              "(Spare) I2C1_SCL(AF4) 또는 GPIO", "여유", FILL_SPARE],
            [53, "PB7",  "—",           "—",              "(Spare) I2C1_SDA(AF4) 또는 GPIO", "여유", FILL_SPARE],
            [54, "BOOT0","—",           "—",              "부트 모드 선택 (점퍼/버튼 권장)", "시스템", FILL_DBG],
            [55, "PB8",  "I2C1_SCL",    "AF4",            "(확장) I2C EEPROM SCL", "I2C 확장", FILL_COMM],
            [56, "PB9",  "I2C1_SDA",    "AF4",            "(확장) I2C EEPROM SDA", "I2C 확장", FILL_COMM],
            [57, "VSS",  "—",           "Power",          "GND", "전원", FILL_PWR],
            [58, "VDD",  "—",           "Power",          "3.3V", "전원", FILL_PWR],
            [59, "VSSA", "—",           "Power",          "VSSA (아날로그 GND)", "전원", FILL_PWR],
            [60, "VREF+","—",           "Power",          "VREF+ (ADC 레퍼런스, 보통 VDDA 연결)", "전원", FILL_PWR],
            [61, "VDDA", "—",           "Power",          "VDDA (아날로그 3.3V, 100nF+1μF 바이패스)", "전원", FILL_PWR],
            [62, "—",    "—",           "Power",          "VSS/VDD (전원)", "전원", FILL_PWR],
            [63, "—",    "—",           "Power",          "VSS/VDD (전원)", "전원", FILL_PWR],
            [64, "—",    "—",           "Power",          "VSS/VDD (전원)", "전원", FILL_PWR],
        ]),
    ]

    headers = ["물리핀#", "포트", "대체 기능(AF)", "모드 설정", "기능 설명", "기능 그룹", "색상구분"]
    r = 3

    for side_name, pins in sides:
        write_section_header(ws, r, side_name, col_count)
        r += 1
        write_header_row(ws, r, headers)
        r += 1
        for pin_data in pins:
            pin_fill = pin_data[6]
            fn = FONT_SPARE if pin_data[5] == "여유" else FONT_NORMAL
            for i in range(6):
                cell = apply_style(ws, r, i+1, fn, pin_fill, ALIGN_CENTER if i != 4 else ALIGN_LEFT)
                cell.value = pin_data[i]
            # 색상 구분 열은 배경색으로 표시
            cell = apply_style(ws, r, 7, fn, pin_fill, ALIGN_CENTER)
            cell.value = pin_data[5]
            r += 1
        r += 1

    # 핀 사용률 통계
    write_section_header(ws, r, "📊 핀 사용률 통계", col_count)
    r += 1
    stats = [
        ["전체 GPIO 핀", 51], ["사용 중 GPIO 핀", 37], ["여유(Spare) 핀", 7],
        ["전원/GND 핀", "~10"], ["디버그 전용(SWD)", 2], ["사용률", "약 73%"],
    ]
    stat_headers = ["항목", "수량", "", "", "", "", ""]
    write_header_row(ws, r, stat_headers)
    r += 1
    for sd in stats:
        cell = apply_style(ws, r, 1, FONT_NORMAL, FILL_WHITE, ALIGN_LEFT)
        cell.value = sd[0]
        cell = apply_style(ws, r, 2, FONT_GROUP if sd[0] == "사용률" else FONT_NORMAL, FILL_WHITE, ALIGN_CENTER)
        cell.value = sd[1]
        r += 1

    ws.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════
# 시트 3: PCB 배치 가이드
# ═══════════════════════════════════════════════════════════
def create_sheet_pcb(wb):
    ws = wb.create_sheet("PCB 배치 가이드")
    ws.sheet_properties.tabColor = "FF6600"
    col_count = 5
    auto_width(ws, [16, 36, 36, 20, 36])

    write_title(ws, 1, "PCB 아트웍 배치 가이드 (LQFP64 기준)", col_count)
    r = 3

    # IC 4면 다이어그램
    write_section_header(ws, r, "🔲 LQFP64 4면 기능 배치 다이어그램", col_count)
    r += 1
    diagram_lines = [
        "              ┌──── Top (49~64) ──────────────────┐",
        "              │ PB3(SWO)  PB4~7(여유/I2C확장)     │",
        "              │ BOOT0  PB8(SCL) PB9(SDA)          │",
        "   Left       │                                    │  Right",
        "  (1~16)      │        STM32G474RCT6               │  (33~48)",
        "  버튼 4ea    │          LQFP64                     │  LCD D7",
        "  ADC(내장VR) │                                    │  PWM CH1",
        "  COMP1(V위상)│                                    │  USART1(디버그)",
        "  USART2 TX/RX│                                    │  SWD(ST-Link)",
        "  ADC(전류)   │                                    │  릴레이 2ea+LED",
        "              └──── Bottom (17~32) ────────────────┘",
        "                RS485 DE/RE   COMP2(I위상)   부저",
        "                ATtiny PWM  외부스위치 3ea  외부VR(ADC)",
        "                LCD RS~D6(6핀연속)   TIM1_CH1N",
    ]
    for line in diagram_lines:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=col_count)
        cell = apply_style(ws, r, 1, Font(name="Consolas", size=9), FILL_WHITE, ALIGN_LEFT)
        cell.value = line
        r += 1
    r += 1

    # 부품 배치 권장
    write_section_header(ws, r, "📐 PCB 영역별 부품 배치 권장", col_count)
    r += 1
    headers = ["PCB 영역", "배치 부품", "MCU 핀 (물리핀#)", "케이블 방향", "비고"]
    write_header_row(ws, r, headers)
    r += 1
    placement = [
        ["좌측 상단", "택트 스위치 4개 + 내장 가변저항", "PC0~PC3(8–11), PA0(12)", "좌측 ←", "사용자 조작부 집중"],
        ["좌측 하단", "MAX3485 + RS485 커넥터", "PA2(14)TX, PA3(15)RX\nPA5(17)DE, PA6(18)/RE", "좌하 ↙", "USART2 인접 최단배선"],
        ["하측 좌", "COMP 아날로그 회로 (CT, 분압기)", "PA1(13)COMP1, PA7(19)COMP2", "하측 ↓", "아날로그 신호 최단 경로"],
        ["하측 중앙", "외부 스위치 커넥터, 외부 가변저항", "PB1~PB2(23–24)\nPB10~PB11(25–26)", "하측 ↓", "외부 케이블 커넥터"],
        ["하측 우", "LCD1602 FPC/핀헤더 커넥터", "PB12~PB15(27–30)\nPC6~PC7(31–32)", "하측 ↓", "6핀 연속 블록"],
        ["우하 코너", "FET 게이트 드라이버 (IR2110 등)", "PB13(28)CH1N\nPA8(35)CH1", "우하 ↘", "고전압부 격리 배치"],
        ["우측 중앙", "ST-Link SWD 커넥터", "PA13(40)SWDIO\nPA14(43)SWCLK", "우측 →", "디버그 커넥터"],
        ["우측", "릴레이 출력 터미널, LED", "PA12(39)Fault\nPA15(44)Running\nPC10(45)LED", "우측 →", "출력 단자대"],
        ["상측", "BOOT0 점퍼, I2C EEPROM (확장)", "BOOT0(54)\nPB8(55)SCL, PB9(56)SDA", "상측 ↑", "점퍼 접근 용이"],
    ]
    for pd in placement:
        fill = FILL_ZEBRA if r % 2 == 0 else FILL_WHITE
        for i, val in enumerate(pd):
            al = ALIGN_LEFT if i >= 1 else ALIGN_CENTER
            cell = apply_style(ws, r, i+1, FONT_NORMAL, fill, al)
            cell.value = val
        r += 1
    r += 1

    # 아트웍 주의사항
    write_section_header(ws, r, "⚠️ PCB 아트웍 주의사항 (필독)", col_count)
    r += 1
    warnings = [
        ["주의사항 1", "PB13(pin28) PWM ↔ LCD 크로스토크",
         "PB13(TIM1_CH1N)이 LCD 핀 블록(PB12~PC7) 사이에 끼어 있음",
         "PWM 트레이스를 별도 레이어로 분리하거나 GND 가드 트레이스 삽입",
         "고전압 PWM 스위칭 노이즈가 LCD 데이터에 간섭 → 문자 깨짐 가능"],
        ["주의사항 2", "PA7(pin19) 아날로그 ↔ PA5/PA6 디지털",
         "PA7(COMP2, 고주파 아날로그)이\nPA5(RS485 DE)·PA6(RS485 /RE)\n디지털 핀과 1핀 간격 인접",
         "① PA6↔PA7 사이 GND 가드 트레이스/비아 펜스\n② COMP2 배선과 RS485 배선 별도 레이어\n③ PA7 하부 연속 GND 플레인 확보\n④ CT→PA7 경로 최단 라우팅, 디지털 교차 금지",
         "RS485 DE/RE 토글 시 고주파 노이즈가\nCOMP2 아날로그 입력에 커플링\n→ PLL 위상 검출 오류 가능"],
        ["주의사항 3", "아날로그 GND 분리",
         "VSSA/VDDA 핀(pin59~61) 영역",
         "아날로그 GND와 디지털 GND를 1점 결합(Star Ground)\nVDDA에 100nF+1μF 바이패스 필수",
         "ADC/COMP 정밀도에 직접 영향"],
        ["주의사항 4", "고전압 PWM 격리",
         "PA8(pin35)·PB13(pin28) 초음파 PWM\n→ IR2110 → 하프브리지 FET",
         "고전압부(파워 스테이지)를 MCU 반대편에 배치\n전력 GND와 신호 GND 분리",
         "FET 스위칭 서지가 MCU에 역류 → 래치업/리셋"],
    ]
    w_headers = ["번호", "항목", "문제 상황", "해결 방안", "미해결 시 위험"]
    write_header_row(ws, r, w_headers)
    r += 1
    for wd in warnings:
        for i, val in enumerate(wd):
            cell = apply_style(ws, r, i+1, FONT_NORMAL if i > 0 else FONT_WARN, FILL_WARN if "주의" in str(wd[0]) else FILL_WHITE, ALIGN_LEFT)
            cell.value = val
        ws.row_dimensions[r].height = 60
        r += 1

    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════
# 시트 4: 핀 변경 이력
# ═══════════════════════════════════════════════════════════
def create_sheet_history(wb):
    ws = wb.create_sheet("핀 변경 이력")
    ws.sheet_properties.tabColor = "7030A0"
    col_count = 6
    auto_width(ws, [10, 22, 12, 12, 50, 14])

    write_title(ws, 1, "핀 변경 이력 (Version History)", col_count)
    r = 3

    # v1.0 → v2.0 변경
    write_section_header(ws, r, "📋 v1.0 → v2.0 핀 재배치 (핵심 아날로그 블록 최적화)", col_count)
    r += 1
    headers = ["버전", "변경 내용", "이전 핀", "현재 핀", "변경 사유", "영향 범위"]
    write_header_row(ws, r, headers)
    r += 1
    changes = [
        ["v2.0", "외부 가변저항 ADC", "PA1", "PB11 (ADC1_IN14)", "PA1을 COMP1_INP 전용 확보\n→ ADC 스캔이 COMP 비교기 입력에 간섭하는 문제 해소", "adc_control, config.h"],
        ["v2.0", "RS485 DE", "PC9", "PA5", "USART2(PA2/PA3) 인접 배치\n→ MAX3485 PCB 배선 최단 거리 확보", "modbus_rtu, config.h"],
        ["v2.0", "RS485 /RE", "PC10", "PA6", "USART2(PA2/PA3) 인접 배치\n→ MAX3485 PCB 배선 최단 거리 확보", "modbus_rtu, config.h"],
        ["v2.0", "Fault 릴레이", "PA6", "PA12", "PA6을 RS485 /RE로 재할당\n→ 릴레이는 속도 무관, 여유 핀으로 이동", "gpio_init, config.h"],
        ["v2.0", "Running 릴레이", "PA7", "PA15", "PA7을 COMP2_INP 전용 확보\n→ 아날로그 모드와 디지털 출력 동시 불가 (핀 충돌 해소)", "gpio_init, config.h"],
        ["v2.0", "상태 LED", "PA5", "PC10", "PA5를 RS485 DE로 재할당\n→ LED는 속도 무관, 여유 핀으로 이동", "gpio_init, config.h"],
    ]
    for cd in changes:
        fill = FILL_ZEBRA if r % 2 == 0 else FILL_WHITE
        for i, val in enumerate(cd):
            al = ALIGN_LEFT if i >= 3 else ALIGN_CENTER
            cell = apply_style(ws, r, i+1, FONT_NORMAL, fill, al)
            cell.value = val
        ws.row_dimensions[r].height = 40
        r += 1
    r += 1

    # v2.0 신규 추가 (내부 아날로그)
    write_section_header(ws, r, "🆕 v2.0 신규 추가 (내부 아날로그 블록 경로)", col_count)
    r += 1
    write_header_row(ws, r, headers)
    r += 1
    new_items = [
        ["v2.0", "COMP1 INM 기준전압", "미지정", "DAC3_CH1 (내부)", "COMP1 반전입력 기준전압\n→ DAC3 내부 전용 DAC로 ~1.65V 공급\n→ 외부 저항 래더 제거, 런타임 보정 가능", "comp_init 신규"],
        ["v2.0", "COMP2 INM 기준전압", "미지정", "DAC3_CH2 (내부)", "COMP2 반전입력 기준전압\n→ DAC3 내부 전용 DAC로 ~1.65V 공급", "comp_init 신규"],
        ["v2.0", "COMP1 출력 → TIM2", "미지정", "TIM2_IC1 (내부)", "전압 제로크로싱 타임스탬프 캡처\n→ 내부 라우팅으로 CPU 개입 없이 캡처", "tim2_ic_init 신규"],
        ["v2.0", "COMP2 출력 → TIM2", "미지정", "TIM2_IC2 (내부)", "전류 제로크로싱 타임스탬프 캡처\n→ 위상차(ns) = (t₂−t₁) × 5.88ns @170MHz", "tim2_ic_init 신규"],
    ]
    for cd in new_items:
        fill = FILL_OK
        for i, val in enumerate(cd):
            al = ALIGN_LEFT if i >= 3 else ALIGN_CENTER
            cell = apply_style(ws, r, i+1, FONT_OK if i == 0 else FONT_NORMAL, fill, al)
            cell.value = val
        ws.row_dimensions[r].height = 50
        r += 1
    r += 1

    # 위상차 측정 정밀도
    write_section_header(ws, r, "📏 위상차 측정 정밀도 (TIM2 듀얼 캡처)", col_count)
    r += 1
    prec_headers = ["동작 주파수", "주기", "1° 대응 시간", "1° 대응 틱 수", "분해능 (°/tick)", "비고"]
    write_header_row(ws, r, prec_headers)
    r += 1
    prec_data = [
        ["28 kHz",  "35.71 μs", "99.2 ns",  "16.9 tick", "0.059°", "가장 높은 정밀도"],
        ["40 kHz",  "25.00 μs", "69.4 ns",  "11.8 tick", "0.085°", "표준 세척 주파수"],
        ["68 kHz",  "14.71 μs", "40.9 ns",  "6.9 tick",  "0.145°", ""],
        ["132 kHz", "7.58 μs",  "21.0 ns",  "3.6 tick",  "0.280°", ""],
        ["168 kHz", "5.95 μs",  "16.5 ns",  "2.8 tick",  "0.356°", "최고 주파수 — 정밀도 한계"],
    ]
    for pd in prec_data:
        fill = FILL_ZEBRA if r % 2 == 0 else FILL_WHITE
        write_data_row(ws, r, pd, fill=fill)
        r += 1
    r += 1

    # 문서 동기화 안내
    write_section_header(ws, r, "📄 문서 동기화 체크리스트", col_count)
    r += 1
    sync_headers = ["파일", "동기화 상태", "마지막 업데이트", "설명", "", ""]
    write_header_row(ws, r, sync_headers)
    r += 1
    sync_data = [
        [".github/copilot-instructions.md", "✅ 동기화 완료", "2026-03-30", "핀맵/COMP/DAC/OPAMP/TIM2 경로 추가", "", ""],
        ["docs/Modbus_레지스터맵.md", "✅ 동기화 완료", "2026-03-30", "RS485 핀 PA5/PA6으로 업데이트", "", ""],
        ["docs/STM32G474RC_핀맵.md", "✅ 동기화 완료", "2026-03-30", "전체 핀맵 md 버전 v2.0 생성", "", ""],
        ["docs/STM32G474RC_핀맵.xlsx", "✅ 동기화 완료", "2026-03-30", "본 파일 (4시트 Excel 버전 v2.0)", "", ""],
        ["include/config.h", "⚠️ 코드 반영 필요", "—", "핀 매크로 정의 업데이트 필요", "", ""],
    ]
    for sd in sync_data:
        fill = FILL_OK if "✅" in str(sd[1]) else FILL_WARN
        for i, val in enumerate(sd):
            cell = apply_style(ws, r, i+1, FONT_NORMAL, fill, ALIGN_LEFT if i in [0,3] else ALIGN_CENTER)
            cell.value = val
        r += 1

    ws.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════
# 메인 실행
# ═══════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    create_sheet_functional(wb)
    create_sheet_physical(wb)
    create_sheet_pcb(wb)
    create_sheet_history(wb)

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "STM32G474RC_핀맵.xlsx")
    wb.save(out_path)
    print(f"✅ 생성 완료: {out_path}")
    print(f"   시트 수: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"   - {name}")


if __name__ == "__main__":
    main()
